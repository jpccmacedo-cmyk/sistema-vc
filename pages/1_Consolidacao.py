from io import BytesIO
from datetime import datetime
import inspect

import streamlit as st
from openpyxl import load_workbook, Workbook

try:
    from utils.ui_vc import (
        configurar_pagina,
        aplicar_css_global,
        render_header,
        render_sidebar_logo,
        render_footer,
        section_title,
    )
except Exception:
    def configurar_pagina(titulo, icone=":bar_chart:", layout="wide"):
        st.set_page_config(page_title=titulo, page_icon=icone, layout=layout)

    def aplicar_css_global(max_width="100%", esconder_sidebar=False):
        st.markdown(
            f"""
            <style>
                .block-container {{max-width:{max_width}!important; padding-top:1.2rem;}}
                .stApp {{background:#F6F8FB;}}
            </style>
            """,
            unsafe_allow_html=True,
        )

    def render_header(titulo, subtitulo=""):
        st.title(titulo)
        if subtitulo:
            st.caption(subtitulo)

    def render_sidebar_logo(texto="Regional Centro-Norte"):
        st.sidebar.caption(texto)

    def render_footer():
        st.caption("Votorantim Cimentos | Regional Centro-Norte")

    def section_title(texto):
        st.subheader(texto)

try:
    import database.consolidados_db as consolidados_db
except Exception:
    consolidados_db = None


configurar_pagina(
    titulo="Consolidação | Sistema CN",
    icone=":inbox_tray:",
    layout="wide",
)
aplicar_css_global(max_width="100%")
render_sidebar_logo()
render_header(
    titulo="Consolidação",
    subtitulo="Upload e processamento dos arquivos mensais | Regional Centro-Norte",
)


def chamar_funcao_flexivel(func, **kwargs):
    assinatura = inspect.signature(func)
    params = assinatura.parameters

    if any(p.kind == inspect.Parameter.VAR_KEYWORD for p in params.values()):
        return func(**kwargs)

    filtrado = {k: v for k, v in kwargs.items() if k in params}

    try:
        return func(**filtrado)
    except TypeError:
        # Tenta chamadas posicionais comuns, para compatibilidade com versões antigas.
        tentativas = [
            [kwargs.get("nome_arquivo"), kwargs.get("arquivo_bytes"), kwargs.get("ano"), kwargs.get("mes")],
            [kwargs.get("nome_arquivo"), kwargs.get("arquivo_bytes"), kwargs.get("ano"), kwargs.get("mes"), kwargs.get("data_processada")],
            [kwargs.get("arquivo_bytes"), kwargs.get("nome_arquivo"), kwargs.get("ano"), kwargs.get("mes")],
        ]
        ultimo_erro = None
        for args in tentativas:
            try:
                return func(*args)
            except TypeError as erro:
                ultimo_erro = erro
        raise ultimo_erro


def inicializar_banco_consolidados():
    if consolidados_db is None:
        return False, "Módulo database.consolidados_db não encontrado."

    init_func = getattr(consolidados_db, "init_consolidados_db", None)
    if callable(init_func):
        init_func()
    return True, "Banco inicializado."


def salvar_consolidado_no_banco(nome_arquivo, arquivo_bytes, ano, mes, data_processada=None):
    if consolidados_db is None:
        return False, "Módulo database.consolidados_db não encontrado."

    nomes_possiveis = [
        "salvar_consolidado",
        "salvar_arquivo_consolidado",
        "criar_consolidado",
        "inserir_consolidado",
        "registrar_consolidado",
    ]

    for nome_funcao in nomes_possiveis:
        func = getattr(consolidados_db, nome_funcao, None)
        if callable(func):
            chamar_funcao_flexivel(
                func,
                nome_arquivo=nome_arquivo,
                arquivo_bytes=arquivo_bytes,
                ano=int(ano),
                mes=int(mes),
                data_processada=data_processada,
                data_geracao=datetime.now(),
            )
            return True, f"Consolidado salvo usando {nome_funcao}."

    return False, "Nenhuma função de salvamento encontrada em database.consolidados_db."


def copiar_aba(origem_ws, destino_wb, nome_aba):
    destino_ws = destino_wb.create_sheet(title=nome_aba[:31])

    for row in origem_ws.iter_rows():
        for cell in row:
            destino_ws[cell.coordinate].value = cell.value

            if cell.has_style:
                destino_ws[cell.coordinate]._style = cell._style
                destino_ws[cell.coordinate].number_format = cell.number_format
                destino_ws[cell.coordinate].font = cell.font.copy()
                destino_ws[cell.coordinate].fill = cell.fill.copy()
                destino_ws[cell.coordinate].border = cell.border.copy()
                destino_ws[cell.coordinate].alignment = cell.alignment.copy()

    for col_letter, col_dim in origem_ws.column_dimensions.items():
        destino_ws.column_dimensions[col_letter].width = col_dim.width

    for row_idx, row_dim in origem_ws.row_dimensions.items():
        destino_ws.row_dimensions[row_idx].height = row_dim.height

    return destino_ws


def gerar_nome_aba_unico(wb, base):
    base = str(base)[:25] if base else "Aba"
    nome = base[:31]
    contador = 1
    while nome in wb.sheetnames:
        sufixo = f"_{contador}"
        nome = (base[:31 - len(sufixo)] + sufixo)[:31]
        contador += 1
    return nome


def consolidar_workbooks(arquivos):
    if len(arquivos) == 1:
        arquivo = arquivos[0]
        return arquivo.name, arquivo.getvalue()

    wb_saida = Workbook()
    wb_saida.remove(wb_saida.active)

    for arquivo in arquivos:
        wb = load_workbook(BytesIO(arquivo.getvalue()), data_only=False)
        prefixo = arquivo.name.rsplit(".", 1)[0]
        for ws in wb.worksheets:
            nome_aba = gerar_nome_aba_unico(wb_saida, f"{prefixo}_{ws.title}")
            copiar_aba(ws, wb_saida, nome_aba)

    output = BytesIO()
    wb_saida.save(output)
    output.seek(0)
    nome_saida = "consolidado_CN.xlsx"
    return nome_saida, output.getvalue()


ok_banco, msg_banco = inicializar_banco_consolidados()
if not ok_banco:
    st.warning(msg_banco)

section_title("Dados do consolidado")

col1, col2, col3 = st.columns([1, 1, 2])
with col1:
    ano = st.number_input("Ano", min_value=2020, max_value=2100, value=datetime.now().year, step=1)
with col2:
    mes = st.number_input("Mês", min_value=1, max_value=12, value=datetime.now().month, step=1)
with col3:
    data_processada = st.date_input("Data do consolidado", value=datetime.now().date())

section_title("Upload")
arquivos = st.file_uploader(
    "Selecione o arquivo consolidado ou os arquivos para consolidar",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True,
)

st.caption(
    "Se enviar apenas um arquivo, ele será salvo como consolidado. Se enviar mais de um, as abas serão reunidas em um único arquivo Excel."
)

if arquivos:
    st.success(f"{len(arquivos)} arquivo(s) carregado(s).")

    with st.expander("Arquivos selecionados", expanded=False):
        for arquivo in arquivos:
            st.write("- " + arquivo.name)

    if st.button("Gerar consolidado", type="primary"):
        try:
            nome_saida, arquivo_saida_bytes = consolidar_workbooks(arquivos)
            st.session_state["consolidado_nome"] = nome_saida
            st.session_state["consolidado_bytes"] = arquivo_saida_bytes
            st.success("Consolidado gerado com sucesso.")
        except Exception as erro:
            st.error("Erro ao gerar consolidado.")
            st.exception(erro)

if "consolidado_bytes" in st.session_state:
    nome_saida = st.session_state["consolidado_nome"]
    arquivo_saida_bytes = st.session_state["consolidado_bytes"]

    section_title("Resultado")

    col_a, col_b = st.columns([1, 1])
    with col_a:
        st.download_button(
            "Baixar consolidado",
            data=arquivo_saida_bytes,
            file_name=nome_saida,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with col_b:
        if st.button("Salvar no histórico compartilhado"):
            try:
                salvo, mensagem = salvar_consolidado_no_banco(
                    nome_arquivo=nome_saida,
                    arquivo_bytes=arquivo_saida_bytes,
                    ano=int(ano),
                    mes=int(mes),
                    data_processada=data_processada,
                )
                if salvo:
                    st.success(mensagem)
                else:
                    st.warning(mensagem)
            except Exception as erro:
                st.error("Erro ao salvar no histórico compartilhado.")
                st.exception(erro)
else:
    st.info("Envie um ou mais arquivos Excel para iniciar a consolidação.")

render_footer()
