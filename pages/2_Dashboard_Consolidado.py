from io import BytesIO

import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

from database.consolidados_db import (
    init_consolidados_db,
    listar_consolidados,
    carregar_arquivo_consolidado,
    excluir_consolidado,
)
from database.metas_db import init_metas_db, carregar_metas
from utils.mapa_indicadores import extrair_resultados_consolidado

from utils.ui_vc import (
    configurar_pagina,
    aplicar_css_global,
    render_header,
    render_sidebar_logo,
    render_footer,
)

try:
    from utils.auth import eh_admin
except Exception:
    def eh_admin():
        return True

configurar_pagina(
    titulo="Dashboard Consolidado | Sistema CN",
    icone=":bar_chart:",
    layout="wide",
)
aplicar_css_global(max_width="100%")
render_sidebar_logo()

PLANTAS_ORDEM = [
    ("COB", "Corumba"),
    ("CUI", "Cuiaba"),
    ("EDE", "Edealina"),
    ("NOB", "Nobres"),
    ("PVE", "Porto Velho"),
    ("SOB", "Sobradinho"),
    ("XAM", "Xambioa"),
    ("CN", "Regional CN"),
]

PLANTAS_STATUS = ["COB", "CUI", "EDE", "NOB", "PVE", "SOB", "XAM"]

GRUPOS_ORDEM = [
    "Fornos",
    "Moagens Cru",
    "Moagens Cimento",
    "Ensacadeiras",
    "Britagens",
    "Estoques",
    "Volumes",
]

INDICADORES_ORDEM = {
    "Fornos": ["OEE", "FP", "FF", "MTBF", "%ST", "CT"],
    "Moagens Cru": ["OEE", "FP", "FF", "MTBF"],
    "Moagens Cimento": ["OEE", "FP", "FF", "MTBF", "%KKC"],
    "Ensacadeiras": ["OEE"],
    "Britagens": ["OEE"],
    "Estoques": ["Clinquer", "Clínquer", "Granel", "Ensacado", "Argamassa"],
    "Volumes": ["Cimento", "Clinquer", "Clínquer"],
}


def to_float(valor):
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except Exception:
        pass
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        texto = valor.strip()
        if texto.upper() in ["", "-", "NA", "N/A", "NAO TEM", "NÃO TEM"]:
            return None
        texto = texto.replace("%", "").replace(".", "").replace(",", ".")
        try:
            return float(texto)
        except Exception:
            return None
    return None


def ajustar_resultado(indicador, valor):
    numero = to_float(valor)
    if numero is None:
        return None
    if indicador in ["%KKC", "KKC"] and abs(numero) <= 1:
        return numero * 100
    return numero


def formatar_numero(valor, tipo=None):
    numero = to_float(valor)
    if numero is None:
        return "NA"
    if tipo == "inteiro":
        return f"{numero:,.0f}".replace(",", ".")
    return f"{numero:.1f}".replace(".", ",")


def formatar_data_farol(consolidado, ano, mes):
    try:
        data_processada = consolidado.get("data_processada")
        if data_processada is not None:
            data = pd.to_datetime(data_processada, errors="coerce")
            if pd.notna(data):
                return data.strftime("%d/%m/%Y")
    except Exception:
        pass
    return "01/" + str(int(mes)).zfill(2) + "/" + str(int(ano))


def calcular_status(resultado, meta, sentido):
    if resultado is None:
        return "sem_resultado"
    if meta is None:
        return "sem_meta"
    if sentido == "informativo":
        return "informativo"
    if sentido == "menor":
        return "verde" if resultado <= meta else "vermelho"
    return "verde" if resultado >= meta else "vermelho"


def css_status(status):
    if status == "verde":
        return "background:#00B050;color:#FFFFFF;font-weight:800;"
    if status == "vermelho":
        return "background:#FF0000;color:#FFFFFF;font-weight:800;"
    if status == "informativo":
        return "background:#D9D9D9;color:#000000;font-weight:800;"
    if status == "sem_meta":
        return "background:#EFEFEF;color:#666666;font-weight:700;"
    return "background:#C9C9C9;color:#000000;font-weight:800;"


def preparar_metas(df_metas):
    df = df_metas.copy()
    for col in ["ano", "mes"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["meta"] = pd.to_numeric(df["meta"], errors="coerce")
    for col in ["codigo", "grupo", "indicador", "sentido", "tipo", "nome", "periodicidade"]:
        if col in df.columns:
            df[col] = df[col].astype(str)
    return df


def buscar_meta(df_metas, ano, mes, codigo, grupo, indicador):
    if df_metas.empty:
        return None
    mensal = df_metas[
        (df_metas["ano"] == int(ano))
        & (df_metas["mes"] == int(mes))
        & (df_metas["codigo"] == codigo)
        & (df_metas["grupo"] == grupo)
        & (df_metas["indicador"] == indicador)
    ]
    if not mensal.empty:
        return mensal.iloc[0]
    anual = df_metas[
        (df_metas["ano"] == int(ano))
        & (df_metas["mes"] == 0)
        & (df_metas["codigo"] == codigo)
        & (df_metas["grupo"] == grupo)
        & (df_metas["indicador"] == indicador)
    ]
    if not anual.empty:
        return anual.iloc[0]
    return None


def montar_base_farol(df_resultados, df_metas, ano, mes):
    df = df_resultados.copy()
    df = df.rename(columns={
        "Planta": "codigo",
        "Grupo": "grupo",
        "Indicador": "indicador",
        "Resultado": "resultado_original",
        "Celula": "celula",
    })
    registros = []
    for _, row in df.iterrows():
        codigo = row["codigo"]
        grupo = row["grupo"]
        indicador = row["indicador"]
        resultado = ajustar_resultado(indicador, row.get("resultado_original"))
        meta_row = buscar_meta(df_metas, ano, mes, codigo, grupo, indicador)
        if meta_row is None:
            meta = None
            sentido = None
            tipo = None
            nome = dict(PLANTAS_ORDEM).get(codigo, codigo)
            periodicidade = None
        else:
            meta = to_float(meta_row.get("meta"))
            sentido = meta_row.get("sentido")
            tipo = meta_row.get("tipo")
            nome = meta_row.get("nome")
            periodicidade = meta_row.get("periodicidade")
        status = calcular_status(resultado, meta, sentido)
        registros.append({
            "Ano": int(ano),
            "Mes": int(mes),
            "Codigo": codigo,
            "Nome": nome,
            "Grupo": grupo,
            "Indicador": indicador,
            "Resultado": resultado,
            "Meta": meta,
            "Sentido": sentido,
            "Tipo": tipo,
            "Periodicidade": periodicidade,
            "Status": status,
        })
    return pd.DataFrame(registros)


def pares_grupo_indicador(df_farol, df_metas, ano, mes):
    pares = set()
    for _, row in df_farol.iterrows():
        pares.add((row["Grupo"], row["Indicador"]))
    metas_cn = df_metas[
        (df_metas["ano"] == int(ano))
        & (df_metas["codigo"] == "CN")
        & (df_metas["mes"].isin([0, int(mes)]))
    ]
    for _, row in metas_cn.iterrows():
        pares.add((row["grupo"], row["indicador"]))

    def chave_ordem(par):
        grupo, indicador = par
        grupo_idx = GRUPOS_ORDEM.index(grupo) if grupo in GRUPOS_ORDEM else 99
        lista_indicadores = INDICADORES_ORDEM.get(grupo, [])
        indicador_idx = lista_indicadores.index(indicador) if indicador in lista_indicadores else 99
        return grupo_idx, indicador_idx, grupo, indicador

    return sorted(pares, key=chave_ordem)


def obter_valores_celula(df_farol, df_metas, ano, mes, codigo, grupo, indicador):
    meta_row = buscar_meta(df_metas, ano, mes, codigo, grupo, indicador)
    if meta_row is None:
        meta = None
        tipo = None
    else:
        meta = to_float(meta_row.get("meta"))
        tipo = meta_row.get("tipo")

    if codigo == "CN":
        return meta, None, tipo, "sem_resultado"

    resultado_row = df_farol[
        (df_farol["Codigo"] == codigo)
        & (df_farol["Grupo"] == grupo)
        & (df_farol["Indicador"] == indicador)
    ]

    if resultado_row.empty:
        return meta, None, tipo, "sem_resultado"

    resultado = resultado_row.iloc[0]["Resultado"]
    status = resultado_row.iloc[0]["Status"]
    tipo_resultado = resultado_row.iloc[0].get("Tipo", tipo)
    if tipo_resultado is not None:
        tipo = tipo_resultado

    return meta, resultado, tipo, status


def gerar_farol_html(df_farol, df_metas, ano, mes, data_farol, plantas_exibir=None):
    if plantas_exibir is None:
        plantas_exibir = PLANTAS_ORDEM

    pares = pares_grupo_indicador(df_farol, df_metas, ano, mes)
    total_cols = 1 + len(plantas_exibir) * 2
    css = """
    <style>
      .farol-wrapper {
        background: #ffffff;
        padding: 12px 14px 18px 14px;
        border-radius: 16px;
        overflow-x: auto;
        width: 100%;
        border: 1px solid #D9E2EC;
        box-shadow: 0 8px 22px rgba(24, 68, 120, 0.05);
      }
      .farol-title {
        font-family: Arial, Helvetica, sans-serif;
        font-size: 24px;
        line-height: 1.15;
        font-weight: 900;
        color: #000000;
        text-align: center;
        margin: 0 0 10px 0;
      }
      .farol-legend {
        display: flex;
        gap: 14px;
        flex-wrap: wrap;
        justify-content: center;
        margin: 0 0 12px 0;
        font-family: Arial, Helvetica, sans-serif;
        font-size: 12px;
        color: #344054;
      }
      .farol-legend-item {
        display: inline-flex;
        align-items: center;
        gap: 5px;
      }
      .farol-legend-box {
        width: 14px;
        height: 14px;
        border: 1px solid #AAB6C2;
        display: inline-block;
      }
      table.farol-table {
        border-collapse: collapse;
        font-family: Arial, Helvetica, sans-serif;
        font-size: 13px;
        color: #000000;
        margin: 0 auto;
        background: #ffffff;
      }
      .farol-table th, .farol-table td {
        border: 1px solid #000000;
        padding: 3px 6px;
        text-align: center;
        vertical-align: middle;
        min-width: 74px;
        height: 22px;
        line-height: 1.1;
      }
      .farol-table th.planta {
        font-weight: 900;
        font-size: 14px;
        background: #ffffff;
        color: #000000;
      }
      .farol-table th.subhead {
        font-weight: 500;
        background: #ffffff;
        color: #000000;
      }
      .farol-table td.indicador {
        min-width: 118px;
        text-align: left;
        font-weight: 700;
        background: #ffffff;
      }
      .farol-table tr.grupo-row td {
        background: #BFBFBF;
        font-weight: 900;
        text-align: left;
        height: 22px;
      }
      .farol-table td.meta {
        background: #ffffff;
        color: #000000;
        font-weight: 500;
      }
      .farol-table td.na-cell {
        background: #C9C9C9;
        color: #000000;
        font-weight: 900;
      }
      .farol-table td.blank-gray {
        background: #C9C9C9;
        color: #000000;
        font-weight: 700;
      }
      .copy-note {
        font-family: Arial, Helvetica, sans-serif;
        color: #6B7280;
        font-size: 12px;
        text-align: right;
        margin-top: 6px;
      }
    </style>
    """
    html = css
    html += "<div class='farol-wrapper' id='farol-copiavel'>"
    html += "<div class='farol-title'>Farol de Indicadores Mensais CN - " + data_farol + "</div>"
    html += """
    <div class='farol-legend'>
      <span class='farol-legend-item'><span class='farol-legend-box' style='background:#00B050;'></span>Dentro da meta</span>
      <span class='farol-legend-item'><span class='farol-legend-box' style='background:#FF0000;'></span>Fora da meta</span>
      <span class='farol-legend-item'><span class='farol-legend-box' style='background:#D9D9D9;'></span>Informativo</span>
      <span class='farol-legend-item'><span class='farol-legend-box' style='background:#EFEFEF;'></span>Sem meta</span>
      <span class='farol-legend-item'><span class='farol-legend-box' style='background:#C9C9C9;'></span>Sem resultado</span>
    </div>
    """
    html += "<table class='farol-table'>"
    html += "<thead>"
    html += "<tr>"
    html += "<th style='border-top:1px solid #ffffff;border-left:1px solid #ffffff;background:#ffffff;'></th>"
    for _, nome_planta in plantas_exibir:
        html += "<th class='planta' colspan='2'>" + nome_planta + "</th>"
    html += "</tr>"
    html += "<tr>"
    html += "<th style='border-left:1px solid #ffffff;background:#ffffff;'></th>"
    for _, _nome_planta in plantas_exibir:
        html += "<th class='subhead'>Meta</th><th class='subhead'>MTD</th>"
    html += "</tr>"
    html += "</thead>"
    html += "<tbody>"
    grupo_atual = None
    for grupo, indicador in pares:
        if grupo != grupo_atual:
            grupo_atual = grupo
            if grupo == "Estoques":
                html += "<tr class='grupo-row'><td>Estoques</td>"
                for _, _nome_planta in plantas_exibir:
                    html += "<td>Cap.</td><td>Real</td>"
                html += "</tr>"
            else:
                html += "<tr class='grupo-row'><td>" + str(grupo) + "</td>"
                for _ in range(total_cols - 1):
                    html += "<td></td>"
                html += "</tr>"
        html += "<tr>"
        html += "<td class='indicador'>" + str(indicador) + "</td>"
        for codigo, _nome_planta in plantas_exibir:
            meta, resultado, tipo, status = obter_valores_celula(df_farol, df_metas, ano, mes, codigo, grupo, indicador)
            meta_txt = formatar_numero(meta, tipo)
            resultado_txt = formatar_numero(resultado, tipo)
            if meta_txt == "NA":
                html += "<td class='blank-gray'>NA</td>"
            else:
                html += "<td class='meta'>" + meta_txt + "</td>"
            if resultado_txt == "NA":
                html += "<td class='na-cell'>NA</td>"
            else:
                html += "<td style='" + css_status(status) + "'>" + resultado_txt + "</td>"
        html += "</tr>"
    html += "</tbody></table>"
    html += "<div class='copy-note'>Dica: use o print da area do navegador ou copie o bloco visual do farol.</div>"
    html += "</div>"
    return html

def calcular_status_por_planta(df_farol):
    df = df_farol[
        df_farol["Codigo"].isin(PLANTAS_STATUS)
        & df_farol["Status"].isin(["verde", "vermelho"])
    ].copy()
    if df.empty:
        return pd.DataFrame()
    resumo = df.groupby(["Codigo", "Status"]).size().unstack(fill_value=0).reset_index()
    if "verde" not in resumo.columns:
        resumo["verde"] = 0
    if "vermelho" not in resumo.columns:
        resumo["vermelho"] = 0
    resumo["total"] = resumo["verde"] + resumo["vermelho"]
    resumo["Acima Meta MTD"] = resumo["verde"] / resumo["total"] * 100
    resumo["Abaixo Meta MTD"] = resumo["vermelho"] / resumo["total"] * 100
    resumo["Nome"] = resumo["Codigo"].map(dict(PLANTAS_ORDEM))
    resumo = resumo.sort_values("Acima Meta MTD", ascending=True)
    return resumo


def plot_status_mtd(df_status):
    fig, ax = plt.subplots(figsize=(10.5, 5.8))
    y = range(len(df_status))
    acima = df_status["Acima Meta MTD"]
    abaixo = df_status["Abaixo Meta MTD"]
    labels = df_status["Codigo"]

    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    ax.barh(y, acima, color="#109618", label="Acima Meta MTD", height=0.78)
    ax.barh(y, abaixo, left=acima, color="#D64550", label="Abaixo Meta MTD", height=0.78)

    for i, (a, b) in enumerate(zip(acima, abaixo)):
        if a > 5:
            ax.text(a / 2, i, f"{a:.2f}%".replace(".", ","), va="center", ha="center", color="white", fontsize=11, fontweight="600")
        if b > 5:
            ax.text(a + b / 2, i, f"{b:.2f}%".replace(".", ","), va="center", ha="center", color="white", fontsize=11, fontweight="600")

    ax.set_yticks(list(y))
    ax.set_yticklabels(labels, fontsize=12, fontweight="700")
    ax.set_xlim(0, 100)
    ax.set_xlabel("")
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, 1.10), ncol=2, frameon=False, fontsize=12)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["bottom"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.tick_params(axis="x", bottom=False, labelbottom=False)
    ax.tick_params(axis="y", length=0)
    fig.subplots_adjust(top=0.82, left=0.08, right=0.98, bottom=0.05)
    return fig



def adicionar_desvios(df_farol):
    df = df_farol.copy()
    df["Desvio"] = df.apply(
        lambda row: None if pd.isna(row.get("Resultado")) or pd.isna(row.get("Meta")) else row.get("Resultado") - row.get("Meta"),
        axis=1,
    )
    df["Desvio %"] = df.apply(
        lambda row: None if pd.isna(row.get("Desvio")) or pd.isna(row.get("Meta")) or row.get("Meta") in [0, None] else row.get("Desvio") / abs(row.get("Meta")) * 100,
        axis=1,
    )
    return df


def filtrar_farol(df_farol, plantas, grupos, status):
    df = df_farol.copy()
    if plantas:
        df = df[df["Codigo"].isin(plantas)]
    if grupos:
        df = df[df["Grupo"].isin(grupos)]
    if status:
        df = df[df["Status"].isin(status)]
    return df.copy()


def filtrar_metas_para_farol(df_metas, ano, mes, plantas, grupos):
    if df_metas.empty:
        return df_metas
    df = df_metas.copy()
    df = df[(df["ano"] == int(ano)) & (df["mes"].isin([0, int(mes)]))]
    if plantas:
        df = df[df["codigo"].isin(plantas)]
    if grupos:
        df = df[df["grupo"].isin(grupos)]
    return df.copy()


def resumo_executivo(df_farol):
    total_aval = int(df_farol["Status"].isin(["verde", "vermelho"]).sum())
    verdes = int((df_farol["Status"] == "verde").sum())
    vermelhos = int((df_farol["Status"] == "vermelho").sum())
    sem_meta = int((df_farol["Status"] == "sem_meta").sum())
    sem_resultado = int((df_farol["Status"] == "sem_resultado").sum())
    pct_verde = (verdes / total_aval * 100) if total_aval else 0
    pct_vermelho = (vermelhos / total_aval * 100) if total_aval else 0
    return {
        "total_aval": total_aval,
        "verdes": verdes,
        "vermelhos": vermelhos,
        "sem_meta": sem_meta,
        "sem_resultado": sem_resultado,
        "pct_verde": pct_verde,
        "pct_vermelho": pct_vermelho,
    }


def calcular_status_executivo(score):
    if score >= 80:
        return "Bom"
    if score >= 60:
        return "Atenção"
    return "Crítico"


def status_por_planta_detalhado(df_farol):
    if df_farol.empty:
        return pd.DataFrame()
    base = df_farol[df_farol["Codigo"].isin(PLANTAS_STATUS)].copy()
    if base.empty:
        return pd.DataFrame()
    tabela = base.groupby(["Codigo", "Status"]).size().unstack(fill_value=0).reset_index()
    for col in ["verde", "vermelho", "sem_meta", "sem_resultado", "informativo"]:
        if col not in tabela.columns:
            tabela[col] = 0
    tabela["Total avaliado"] = tabela["verde"] + tabela["vermelho"]
    tabela["Score"] = tabela.apply(lambda row: (row["verde"] / row["Total avaliado"] * 100) if row["Total avaliado"] else 0, axis=1)
    tabela["Status executivo"] = tabela["Score"].apply(calcular_status_executivo)
    tabela["Nome"] = tabela["Codigo"].map(dict(PLANTAS_ORDEM))
    tabela = tabela.rename(columns={
        "verde": "Verde",
        "vermelho": "Vermelho",
        "sem_meta": "Sem meta",
        "sem_resultado": "Sem resultado",
        "informativo": "Informativo",
    })
    return tabela[["Codigo", "Nome", "Verde", "Vermelho", "Sem meta", "Sem resultado", "Informativo", "Total avaliado", "Score", "Status executivo"]].sort_values("Score", ascending=False)


def ranking_indicadores_criticos(df_farol):
    if df_farol.empty:
        return pd.DataFrame()
    criticos = df_farol[df_farol["Status"] == "vermelho"].copy()
    if criticos.empty:
        return pd.DataFrame()
    ranking = criticos.groupby(["Grupo", "Indicador"]).size().reset_index(name="Ocorrências fora da meta")
    return ranking.sort_values("Ocorrências fora da meta", ascending=False)


def indicadores_sem_meta(df_farol):
    cols = ["Codigo", "Nome", "Grupo", "Indicador", "Resultado", "Meta", "Status"]
    return df_farol[df_farol["Status"] == "sem_meta"][cols].copy()


def buscar_consolidado_mes_anterior(df_consolidados, ano, mes):
    ano = int(ano)
    mes = int(mes)
    if mes == 1:
        ano_ant, mes_ant = ano - 1, 12
    else:
        ano_ant, mes_ant = ano, mes - 1
    df_prev = df_consolidados[(df_consolidados["ano"] == ano_ant) & (df_consolidados["mes"] == mes_ant)].copy()
    if df_prev.empty:
        return None, ano_ant, mes_ant
    if "data_geracao" in df_prev.columns:
        df_prev = df_prev.sort_values("data_geracao")
    return df_prev.iloc[-1], ano_ant, mes_ant


def carregar_farol_mes_anterior(df_consolidados, df_metas, ano, mes):
    consolidado_ant, ano_ant, mes_ant = buscar_consolidado_mes_anterior(df_consolidados, ano, mes)
    if consolidado_ant is None:
        return pd.DataFrame(), ano_ant, mes_ant, None
    try:
        nome_ant, bytes_ant = carregar_arquivo_consolidado(consolidado_ant["consolidado_id"])
        if not bytes_ant:
            return pd.DataFrame(), ano_ant, mes_ant, None
        wb_ant = load_workbook(BytesIO(bytes_ant), data_only=True)
        resultados_ant = extrair_resultados_consolidado(wb_ant)
        farol_ant = montar_base_farol(resultados_ant, df_metas, int(ano_ant), int(mes_ant))
        farol_ant = adicionar_desvios(farol_ant)
        return farol_ant, ano_ant, mes_ant, nome_ant
    except Exception:
        return pd.DataFrame(), ano_ant, mes_ant, None


def comparar_com_mes_anterior(df_atual, df_anterior):
    if df_atual.empty or df_anterior.empty:
        return pd.DataFrame()
    chaves = ["Codigo", "Grupo", "Indicador"]
    atual = df_atual[chaves + ["Resultado", "Status"]].rename(columns={"Resultado": "Resultado atual", "Status": "Status atual"})
    anterior = df_anterior[chaves + ["Resultado", "Status"]].rename(columns={"Resultado": "Resultado anterior", "Status": "Status anterior"})
    comp = atual.merge(anterior, on=chaves, how="inner")
    comp["Delta"] = comp["Resultado atual"] - comp["Resultado anterior"]
    return comp.sort_values("Delta", ascending=False)

init_consolidados_db()
init_metas_db()

render_header(
    titulo="Dashboard Consolidado",
    subtitulo="Farol mensal por planta, status MTD e histórico compartilhado | Regional Centro-Norte",
)

try:
    df_consolidados = listar_consolidados()
    df_metas = preparar_metas(carregar_metas())
except Exception as e:
    st.error("Erro ao carregar dados do banco.")
    st.exception(e)
    st.stop()

if df_consolidados.empty:
    st.warning("Nenhum consolidado salvo ainda. Gere um consolidado na pagina de Consolidacao.")
    st.stop()

if df_metas.empty:
    st.warning("Nenhuma meta cadastrada ainda. Faca upload da planilha na pagina Metas.")
    st.stop()

st.sidebar.header("Filtros")
df_consolidados["ano"] = pd.to_numeric(df_consolidados["ano"], errors="coerce")
df_consolidados["mes"] = pd.to_numeric(df_consolidados["mes"], errors="coerce")
anos = sorted(df_consolidados["ano"].dropna().unique())
ano_sel = st.sidebar.selectbox("Ano", anos, index=len(anos) - 1)
df_ano = df_consolidados[df_consolidados["ano"] == ano_sel].copy()
meses = sorted(df_ano["mes"].dropna().unique())
mes_sel = st.sidebar.selectbox("Mes", meses, index=len(meses) - 1)
df_mes = df_ano[df_ano["mes"] == mes_sel].copy()
opcoes = []
for _, row in df_mes.iterrows():
    opcoes.append(str(row["nome_arquivo"]) + " - " + str(row["data_geracao"]))

if not opcoes:
    st.warning("Nenhum consolidado encontrado para os filtros selecionados.")
    st.stop()

opcao = st.sidebar.selectbox("Consolidado", opcoes)
idx = opcoes.index(opcao)
consolidado = df_mes.iloc[idx]
consolidado_id = consolidado["consolidado_id"]
data_farol = formatar_data_farol(consolidado, ano_sel, mes_sel)
nome_arquivo, arquivo_bytes = carregar_arquivo_consolidado(consolidado_id)

if not arquivo_bytes:
    st.error("Arquivo consolidado nao encontrado no banco.")
    st.stop()

try:
    wb = load_workbook(BytesIO(arquivo_bytes), data_only=True)
    df_resultados = extrair_resultados_consolidado(wb)
except Exception as e:
    st.error("Erro ao extrair indicadores do consolidado.")
    st.exception(e)
    st.stop()

if df_resultados.empty:
    st.warning("Nenhum indicador extraido do consolidado.")
    st.stop()

df_farol = montar_base_farol(df_resultados, df_metas, int(ano_sel), int(mes_sel))
df_farol = adicionar_desvios(df_farol)

# Filtros executivos do farol
st.sidebar.header("Filtros do Farol")
plantas_disponiveis = [codigo for codigo, _ in PLANTAS_ORDEM if codigo in set(df_farol["Codigo"].dropna().unique()) or codigo == "CN"]
grupos_disponiveis = [g for g in GRUPOS_ORDEM if g in set(df_farol["Grupo"].dropna().unique())]
status_labels = {
    "verde": "Dentro da meta",
    "vermelho": "Fora da meta",
    "sem_meta": "Sem meta",
    "sem_resultado": "Sem resultado",
    "informativo": "Informativo",
}
status_disponiveis = [s for s in status_labels if s in set(df_farol["Status"].dropna().unique())]

plantas_sel = st.sidebar.multiselect("Plantas", plantas_disponiveis, default=plantas_disponiveis)
grupos_sel = st.sidebar.multiselect("Grupos", grupos_disponiveis, default=grupos_disponiveis)
status_sel_labels = st.sidebar.multiselect(
    "Status",
    [status_labels[s] for s in status_disponiveis],
    default=[status_labels[s] for s in status_disponiveis],
)
status_sel = [s for s, label in status_labels.items() if label in status_sel_labels]

if st.sidebar.button("Limpar filtros do Farol"):
    st.rerun()

df_farol_filtrado = filtrar_farol(df_farol, plantas_sel, grupos_sel, status_sel)
df_metas_farol = filtrar_metas_para_farol(df_metas, int(ano_sel), int(mes_sel), plantas_sel, grupos_sel)
plantas_exibir = [(codigo, nome) for codigo, nome in PLANTAS_ORDEM if codigo in plantas_sel]

if df_farol_filtrado.empty:
    st.warning("Nenhum indicador encontrado com os filtros do farol selecionados.")

resumo = resumo_executivo(df_farol_filtrado)
df_status = calcular_status_por_planta(df_farol_filtrado)
df_status_exec = status_por_planta_detalhado(df_farol_filtrado)
df_criticos = ranking_indicadores_criticos(df_farol_filtrado)
df_sem_meta = indicadores_sem_meta(df_farol_filtrado)
html_farol = gerar_farol_html(df_farol_filtrado, df_metas_farol, int(ano_sel), int(mes_sel), data_farol, plantas_exibir=plantas_exibir)

farol_anterior, ano_ant, mes_ant, nome_ant = carregar_farol_mes_anterior(df_consolidados, df_metas, int(ano_sel), int(mes_sel))
comparativo_mes = comparar_com_mes_anterior(df_farol_filtrado, farol_anterior)

st.subheader("Resumo executivo")
col1, col2, col3, col4, col5 = st.columns(5)
col1.metric("Na meta", f"{resumo['pct_verde']:.1f}%".replace(".", ","), f"{resumo['verdes']} indicadores")
col2.metric("Fora da meta", f"{resumo['pct_vermelho']:.1f}%".replace(".", ","), f"{resumo['vermelhos']} indicadores")
col3.metric("Sem meta", int(resumo["sem_meta"]))
col4.metric("Sem resultado", int(resumo["sem_resultado"]))
col5.metric("Avaliados", int(resumo["total_aval"]))

st.caption("Arquivo em análise: " + str(nome_arquivo) + " | Farol: " + str(data_farol))
st.download_button(
    "Baixar arquivo consolidado",
    data=arquivo_bytes,
    file_name=nome_arquivo,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

if not df_sem_meta.empty:
    with st.expander("Indicadores sem meta cadastrada", expanded=False):
        st.warning(f"Foram encontrados {len(df_sem_meta)} indicador(es) sem meta nos filtros atuais.")
        st.dataframe(df_sem_meta, use_container_width=True, hide_index=True)

st.divider()

tabs_nomes = [
    "Análise Executiva",
    "Farol Mensal",
    "Status por Planta",
    "Indicadores Críticos",
    "Detalhes",
    "Histórico",
]
if eh_admin():
    tabs_nomes.append("Administração")
else:
    st.sidebar.info("Aba Administração oculta para usuário comum.")

tabs = st.tabs(tabs_nomes)
tab_map = dict(zip(tabs_nomes, tabs))

with tab_map["Análise Executiva"]:
    st.subheader("Ranking de plantas")
    if df_status_exec.empty:
        st.info("Nao ha indicadores suficientes para montar o ranking por planta.")
    else:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("#### Melhores plantas")
            melhores = df_status_exec.sort_values("Score", ascending=False).head(5).copy()
            melhores["Score"] = melhores["Score"].map(lambda x: f"{x:.1f}%".replace(".", ","))
            st.dataframe(melhores, use_container_width=True, hide_index=True)
        with c2:
            st.markdown("#### Plantas com maior atenção")
            atencao = df_status_exec.sort_values("Score", ascending=True).head(5).copy()
            atencao["Score"] = atencao["Score"].map(lambda x: f"{x:.1f}%".replace(".", ","))
            st.dataframe(atencao, use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("Comparação com mês anterior")
    if comparativo_mes.empty:
        st.info(f"Nao foi encontrado consolidado comparável para {str(mes_ant).zfill(2)}/{ano_ant}.")
    else:
        st.caption("Comparando com: " + str(nome_ant))
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("#### Maiores evoluções")
            evolucoes = comparativo_mes.sort_values("Delta", ascending=False).head(10).copy()
            st.dataframe(evolucoes, use_container_width=True, hide_index=True)
        with c2:
            st.markdown("#### Maiores quedas")
            quedas = comparativo_mes.sort_values("Delta", ascending=True).head(10).copy()
            st.dataframe(quedas, use_container_width=True, hide_index=True)

with tab_map["Farol Mensal"]:
    st.markdown(html_farol, unsafe_allow_html=True)
    st.download_button(
        "Baixar farol em HTML",
        data=html_farol.encode("utf-8"),
        file_name="farol_mensal_cn.html",
        mime="text/html"
    )

with tab_map["Status por Planta"]:
    st.markdown(
        "<h2 style='text-align:center; color:#1A2A8F; font-weight:900; margin-bottom:18px;'>Status Indicador (MTD)</h2>",
        unsafe_allow_html=True
    )
    st.caption("CN nao e calculado neste grafico. O status considera apenas COB, CUI, EDE, NOB, PVE, SOB e XAM.")
    if df_status.empty:
        st.info("Nao ha indicadores suficientes com meta e resultado para montar o status MTD.")
    else:
        fig = plot_status_mtd(df_status)
        st.pyplot(fig, use_container_width=True)
        df_status_exibir = df_status.copy()
        df_status_exibir["Acima Meta MTD"] = df_status_exibir["Acima Meta MTD"].map(lambda x: f"{x:.2f}%".replace(".", ","))
        df_status_exibir["Abaixo Meta MTD"] = df_status_exibir["Abaixo Meta MTD"].map(lambda x: f"{x:.2f}%".replace(".", ","))
        st.dataframe(df_status_exibir, use_container_width=True, hide_index=True)

    st.markdown("#### Status consolidado por planta")
    if df_status_exec.empty:
        st.info("Sem dados para status consolidado por planta.")
    else:
        tabela_exec = df_status_exec.copy()
        tabela_exec["Score"] = tabela_exec["Score"].map(lambda x: f"{x:.1f}%".replace(".", ","))
        st.dataframe(tabela_exec, use_container_width=True, hide_index=True)

with tab_map["Indicadores Críticos"]:
    st.subheader("Indicadores mais críticos")
    if df_criticos.empty:
        st.success("Nenhum indicador fora da meta nos filtros atuais.")
    else:
        st.dataframe(df_criticos, use_container_width=True, hide_index=True)

    st.subheader("Indicadores sem meta")
    if df_sem_meta.empty:
        st.success("Nenhum indicador sem meta nos filtros atuais.")
    else:
        st.dataframe(df_sem_meta, use_container_width=True, hide_index=True)

with tab_map["Detalhes"]:
    st.subheader("Base detalhada do farol")
    df_view = df_farol_filtrado.copy()
    st.dataframe(df_view, use_container_width=True, hide_index=True)
    csv = df_view.to_csv(index=False).encode("utf-8-sig")
    st.download_button("Baixar detalhes em CSV", data=csv, file_name="farol_consolidado_detalhado.csv", mime="text/csv")

with tab_map["Histórico"]:
    st.subheader("Historico de consolidados")
    cols = ["nome_arquivo", "data_processada", "ano", "mes", "data_geracao"]
    cols = [c for c in cols if c in df_consolidados.columns]
    st.dataframe(df_consolidados[cols], use_container_width=True, hide_index=True)

if "Administração" in tab_map:
    with tab_map["Administração"]:
        st.subheader("Administracao")
        with st.expander("Excluir consolidado selecionado"):
            st.warning("Essa acao remove o consolidado do historico compartilhado.")
            confirmacao = st.text_input("Digite EXCLUIR para confirmar", value="")
            if st.button("Excluir consolidado selecionado", disabled=confirmacao != "EXCLUIR"):
                excluir_consolidado(consolidado_id)
                st.success("Consolidado excluido.")
                st.rerun()

render_footer()
