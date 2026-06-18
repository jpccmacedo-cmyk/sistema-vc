from datetime import datetimefrom datetime import": grupo,
                "indicador": indicador,
                "meta": valor_meta,
                "sentido": obter_sentido(grupo, indicador, tipos_lookup),
                "tipo": tipo_indicador(indicador),
                "observacao": observacao,
            })


def adicionar_meta_anual(
    registros,
    ws,
    tipos_lookup,
    titulo_origem,
    row_start,
    row_end,
    label_col,
    value_col,
    grupo,
    indicador,
    ano=2026
):
    for linha in range(row_start, row_end + 1):
        label = ws.cell(linha, label_col).value
        planta_info = identificar_planta(label)

        if not planta_info:
            continue

        valor_original = limpar_valor(ws.cell(linha, value_col).value)

        if valor_original is None:
            continue

        valor_meta = ajustar_meta(indicador, valor_original)

        nivel, codigo, nome = planta_info

        registros.append({
            "ano": ano,
            "mes": 0,
            "periodicidade": "Anual",
            "nivel": nivel,
            "codigo": codigo,
            "nome": nome,
            "grupo": grupo,
            "indicador": indicador,
            "meta": valor_meta,
            "sentido": obter_sentido(grupo, indicador, tipos_lookup),
            "tipo": tipo_indicador(indicador),
            "observacao": f"Origem: {titulo_origem}",
        })


def normalizar_planilha_metas(caminho_arquivo):
    wb = load_workbook(caminho_arquivo, data_only=True)

    if "METAS 2026" not in wb.sheetnames:
        raise Exception("A aba 'METAS 2026' não foi encontrada.")

    ws = wb["METAS 2026"]
    tipos_lookup = carregar_tipos_planilha(wb)

    registros = []

    # Metas mensais
    adicionar_meta_mensal(
        registros, ws, tipos_lookup,
        "OEE - Fornos",
        4, 6, 13, 3, 4, 15,
        "Fornos", "OEE"
    )

    adicionar_meta_mensal(
        registros, ws, tipos_lookup,
        "OEE - Moagens de Cimento",
        16, 18, 25, 3, 4, 15,
        "Moagens Cimento", "OEE"
    )

    adicionar_meta_mensal(
        registros, ws, tipos_lookup,
        "KKC",
        28, 30, 37, 3, 4, 15,
        "Moagens Cimento", "%KKC"
    )

    adicionar_meta_mensal(
        registros, ws, tipos_lookup,
        "ST",
        40, 42, 49, 3, 4, 15,
        "Fornos", "%ST"
    )

    # Segundo bloco de OEE Moagem Cimento existente na planilha.
    # Será deduplicado pela chave.
    adicionar_meta_mensal(
        registros, ws, tipos_lookup,
        "OEE MOAGEM CIMENTO",
        72, 74, 81, 3, 4, 15,
        "Moagens Cimento", "OEE"
    )

    # Metas anuais
    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "FP Fornos",
        52, 59, 3, 4,
        "Fornos", "FP"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "FF Fornos",
        52, 59, 6, 7,
        "Fornos", "FF"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "MTBF Fornos",
        52, 59, 10, 11,
        "Fornos", "MTBF"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "OEE MOAGEM CRU",
        62, 69, 3, 4,
        "Moagens Cru", "OEE"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "FP MOAGEM CIMENTO",
        84, 91, 3, 4,
        "Moagens Cimento", "FP"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "FF MOAGEM CIMENTO",
        84, 91, 6, 7,
        "Moagens Cimento", "FF"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "MTBF MOAGEM CIMENTO",
        84, 91, 9, 10,
        "Moagens Cimento", "MTBF"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "OEE ENSACADEIRAS",
        94, 101, 3, 4,
        "Ensacadeiras", "OEE"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "OEE BRITAGENS",
        94, 101, 6, 7,
        "Britagens", "OEE"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "OEE MOAGENS CRU",
        94, 101, 9, 10,
        "Moagens Cru", "OEE"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "FP MOAGENS CRU",
        94, 101, 12, 13,
        "Moagens Cru", "FP"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "FF MOAGENS CRU",
        94, 101, 15, 16,
        "Moagens Cru", "FF"
    )

    adicionar_meta_anual(
        registros, ws, tipos_lookup,
        "MTBF MOAGENS CRU",
        94, 101, 18, 19,
        "Moagens Cru", "MTBF"
    )

    df = pd.DataFrame(registros)

    if df.empty:
        return df

    chave = [
        "ano",
        "mes",
        "periodicidade",
        "nivel",
        "codigo",
        "grupo",
        "indicador",
    ]

    df = df.sort_values(chave + ["observacao"])
    df = df.drop_duplicates(subset=chave, keep="first").reset_index(drop=True)

    return df


def salvar_metas_no_banco(df_metas, usuario_upload=None, arquivo_origem=None):
    init_metas_db()

    if df_metas.empty:
        return 0

    data_upload = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    df = df_metas.copy()
    df["id"] = [str(uuid.uuid4()) for _ in range(len(df))]
    df["data_upload"] = data_upload
    df["usuario_upload"] = usuario_upload
    df["arquivo_origem"] = arquivo_origem

    colunas = [
        "id",
        "ano",
        "mes",
        "periodicidade",
        "nivel",
        "codigo",
        "nome",
        "grupo",
        "indicador",
        "meta",
        "sentido",
        "tipo",
        "observacao",
        "data_upload",
        "usuario_upload",
        "arquivo_origem",
    ]

    engine = get_engine()

    with engine.begin() as conn:
        for _, row in df.iterrows():
            conn.execute(
                text("""
                    DELETE FROM metas_consolidados
                    WHERE ano = :ano
                      AND mes = :mes
                      AND codigo = :codigo
                      AND grupo = :grupo
                      AND indicador = :indicador
                """),
                {
                    "ano": int(row["ano"]),
                    "mes": int(row["mes"]),
                    "codigo": row["codigo"],
                    "grupo": row["grupo"],
                    "indicador": row["indicador"],
                }
            )

    df[colunas].to_sql(
        "metas_consolidados",
        engine,
        if_exists="append",
        index=False
    )

    return len(df)


def carregar_metas():
    init_metas_db()

    engine = get_engine()

    query = """
        SELECT *
        FROM metas_consolidados
        ORDER BY ano DESC, mes DESC, nivel, codigo, grupo, indicador
    """

    return pd.read_sql(query, engine)


def excluir_metas_ano(ano):
    init_metas_db()

    engine = get_engine()

    with engine.begin() as conn:
        conn.execute(
            text("DELETE FROM metas_consolidados WHERE ano = :ano"),
            {"ano": int(ano)}
        )
import uuid
import re
import unicodedata

import pandas as pd
from sqlalchemy import text
from openpyxl import load_workbook

from database.connection import get_engine


PLANT_MAP = {
    "centro-norte": ("Regional", "CN", "Regional CN"),
    "corumba": ("Planta", "COB", "Corumbá"),
    "cuiaba": ("Planta", "CUI", "Cuiabá"),
    "edealina": ("Planta", "EDE", "Edealina"),
    "nobres": ("Planta", "NOB", "Nobres"),
    "porto velho": ("Planta", "PVE", "Porto Velho"),
    "sobradinho": ("Planta", "SOB", "Sobradinho"),
    "xambioa": ("Planta", "XAM", "Xambioá"),
}


ALIASES_TIPOS = {
    ("Fornos", "OEE"): "oee fornos",
    ("Fornos", "%ST"): "st",
    ("Fornos", "FP"): "fp fornos",
    ("Fornos", "FF"): "ff fornos",
    ("Fornos", "MTBF"): "mtbf fornos",

    ("Moagens Cimento", "OEE"): "oee moagem cimento",
    ("Moagens Cimento", "%KKC"): "kkc",
    ("Moagens Cimento", "KKC"): "kkc",
    ("Moagens Cimento", "FP"): "fp moagem cimento",
    ("Moagens Cimento", "FF"): "ff moagem cimento",
    ("Moagens Cimento", "MTBF"): "mtbf moagem cimento",

    ("Moagens Cru", "OEE"): "oee moagens cru",
    ("Moagens Cru", "FP"): "fp moagens cru",
    ("Moagens Cru", "FF"): "ff moagens cru",
    ("Moagens Cru", "MTBF"): "mtbf moagens cru",

    ("Ensacadeiras", "OEE"): "oee ensacadeira",
    ("Britagens", "OEE"): "oee britagem",
}


def normalizar_texto(valor):
    texto = "" if valor is None else str(valor)
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ASCII", "ignore").decode("ASCII")
    texto = texto.lower().strip()
    texto = texto.replace("\n", " ")
    texto = re.sub(r"[-_]+", " ", texto)
    texto = re.sub(r"\btotal\b", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def identificar_planta(label):
    chave = normalizar_texto(label)

    for inicio, dados in PLANT_MAP.items():
        if chave.startswith(inicio):
            return dados

    return None


def limpar_valor(valor):
    if valor is None:
        return None

    if isinstance(valor, str):
        texto = valor.strip().upper()

        if texto in ["", "-", "NA", "N/A"]:
            return None

        try:
            return float(texto.replace(",", "."))
        except Exception:
            return None

    try:
        return float(valor)
    except Exception:
        return None


def sentido_from_tipo(tipo):
    texto = normalizar_texto(tipo)

    if "menor" in texto:
        return "menor"

    if "maior" in texto:
        return "maior"

    return "informativo"


def tipo_indicador(indicador):
    if indicador in ["%KKC", "KKC"]:
        return "numero"

    if indicador in ["Clínquer", "Granel", "Ensacado", "Argamassa", "Cimento"]:
        return "inteiro"

    return "numero"


def ajustar_meta(indicador, valor):
    """
    Regra específica do KKC:
    Se vier 0,55 no Excel, salvar como 55.
    Ou seja, KKC fica número puro, sem símbolo de %.
    """

    if indicador in ["%KKC", "KKC"] and valor is not None and abs(valor) <= 1:
        return valor * 100

    return valor


def init_metas_db():
    engine = get_engine()

    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS metas_consolidados (
                id TEXT PRIMARY KEY,
                ano INTEGER NOT NULL,
                mes INTEGER NOT NULL,
                periodicidade TEXT,
                nivel TEXT,
                codigo TEXT,
                nome TEXT,
                grupo TEXT,
                indicador TEXT,
                meta REAL,
                sentido TEXT,
                tipo TEXT,
                observacao TEXT,
                data_upload TEXT,
                usuario_upload TEXT,
                arquivo_origem TEXT
            )
        """))

        conn.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_metas_consolidados_busca
            ON metas_consolidados (ano, mes, codigo, grupo, indicador)
        """))


def carregar_tipos_planilha(wb):
    if "TIPOS" not in wb.sheetnames:
        return {}

    ws = wb["TIPOS"]
    tipos = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        valores = [v for v in row if v is not None]

        if len(valores) >= 2:
            indicador = str(valores[0]).strip()
            tipo_meta = str(valores[1]).strip()
            tipos[normalizar_texto(indicador)] = sentido_from_tipo(tipo_meta)

    return tipos


def obter_sentido(grupo, indicador, tipos_lookup):
    chave = ALIASES_TIPOS.get((grupo, indicador), indicador)
    chave_norm = normalizar_texto(chave)

    return tipos_lookup.get(chave_norm, "maior")


def adicionar_meta_mensal(
    registros,
    ws,
    tipos_lookup,
    titulo_origem,
    header_row,
    first_data_row,
    last_data_row,
    label_col,
    first_month_col,
    last_month_col,
    grupo,
    indicador
):
    for linha in range(first_data_row, last_data_row + 1):
        label = ws.cell(linha, label_col).value
        planta_info = identificar_planta(label)

        if not planta_info:
            continue

        nivel, codigo, nome = planta_info

        for coluna in range(first_month_col, last_month_col + 1):
            data_mes = ws.cell(header_row, coluna).value

            if not isinstance(data_mes, datetime):
                continue

            valor_original = limpar_valor(ws.cell(linha, coluna).value)

            if valor_original is None:
                continue

            valor_meta = ajustar_meta(indicador, valor_original)

            observacao = f"Origem: {titulo_origem}"

            if indicador in ["%KKC", "KKC"]:
                observacao += " | KKC convertido para número sem %. Ex.: 0,55 virou 55."

            registros.append({
                "ano": data_mes.year,
                "mes": data_mes.month,
                "periodicidade": "Mensal",
                "nivel": nivel,
                "codigo": codigo,
                "nome": nome,
